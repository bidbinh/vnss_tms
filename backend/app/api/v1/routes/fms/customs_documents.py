"""
FMS Customs Document Upload & Parsing API Routes
Upload và phân tích chứng từ để tạo tờ khai hải quan
"""
import os
import uuid
import io
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Query
from sqlmodel import Session, select
from typing import Optional, List, Dict, Any
from datetime import datetime
from pydantic import BaseModel, Field
from enum import Enum

from app.db.session import get_session
from app.models import User
from app.core.security import get_current_user
from app.services.document_parser import (
    DocumentParser,
    ParsedDocument,
    ParsedItem,
    parse_pdf_content,
    parse_excel_content,
)
from app.services.ai_document_parser import (
    AIDocumentParser,
    AIParseResult,
    parse_document_with_ai,
)

router = APIRouter(prefix="/customs/documents", tags=["FMS Customs Documents"])


# Configuration
# Use temp directory that works on both Windows and Linux
import tempfile
_default_upload_dir = os.path.join(tempfile.gettempdir(), "uploads", "customs")
UPLOAD_DIR = os.environ.get("UPLOAD_DIR", _default_upload_dir)
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
ALLOWED_EXTENSIONS = {".pdf", ".xlsx", ".xls", ".csv"}


class DocumentTypeEnum(str, Enum):
    INVOICE = "INVOICE"
    PACKING_LIST = "PACKING_LIST"
    BILL_OF_LADING = "BILL_OF_LADING"
    AIRWAY_BILL = "AIRWAY_BILL"
    ARRIVAL_NOTICE = "ARRIVAL_NOTICE"


class ParsedItemResponse(BaseModel):
    item_no: int
    product_code: Optional[str] = None      # Customer PN / DELL PN
    supplier_code: Optional[str] = None     # Supplier PN / Model
    product_name: Optional[str] = None      # Description of Goods
    hs_code: Optional[str] = None
    quantity: float = 0
    unit: Optional[str] = None
    unit_price: float = 0
    total_value: float = 0
    gross_weight: float = 0
    net_weight: float = 0
    country_of_origin: Optional[str] = None


class ParsedDocumentResponse(BaseModel):
    """Response schema for parsed document"""
    document_type: Optional[str]
    confidence: float

    # Header info
    invoice_no: Optional[str]
    invoice_date: Optional[str]
    bl_no: Optional[str]
    bl_date: Optional[str]

    # Parties
    seller_name: Optional[str]
    seller_address: Optional[str]
    consignee_name: Optional[str]
    consignee_address: Optional[str]

    # Transport
    vessel_name: Optional[str]
    voyage_no: Optional[str]
    flight_no: Optional[str]
    loading_port: Optional[str]
    discharge_port: Optional[str]
    eta: Optional[str]

    # Cargo
    total_packages: int
    gross_weight: float
    net_weight: float
    volume: float
    container_numbers: Optional[str]

    # Values
    currency: str
    total_value: float
    incoterms: Optional[str]

    # Items
    items: List[ParsedItemResponse]

    # Metadata
    warnings: List[str]
    extracted_fields: Dict[str, Any]


class UploadResponse(BaseModel):
    """Response for file upload"""
    file_id: str
    file_name: str
    file_size: int
    mime_type: str
    uploaded_at: datetime


class CustomsDataPreview(BaseModel):
    """Preview of customs declaration data from parsed documents"""
    # Declaration fields
    declaration_type: str = "IMPORT"
    foreign_partner_name: Optional[str]
    foreign_partner_address: Optional[str]
    trader_name: Optional[str]
    trader_address: Optional[str]

    invoice_no: Optional[str]
    invoice_date: Optional[str]
    bl_no: Optional[str]
    bl_date: Optional[str]

    vessel_name: Optional[str]
    voyage_no: Optional[str]
    loading_port: Optional[str]
    discharge_port: Optional[str]

    total_packages: int
    gross_weight: float
    net_weight: float
    container_numbers: Optional[str]

    currency_code: str
    fob_value: float
    incoterms: Optional[str]

    # HS Code items
    items: List[ParsedItemResponse]

    # Source info
    source_documents: List[str]
    data_confidence: float


@router.post("/upload", response_model=UploadResponse)
async def upload_document(
    file: UploadFile = File(...),
    document_type: Optional[DocumentTypeEnum] = Form(None),
    current_user: User = Depends(get_current_user),
):
    """
    Upload a document file for parsing.

    Supported formats:
    - PDF: Invoice, Packing List, Bill of Lading, Arrival Notice
    - Excel: Packing List, Item lists

    Returns a file_id that can be used for parsing.
    """
    import logging
    logger = logging.getLogger(__name__)

    # Validate file extension
    file_ext = os.path.splitext(file.filename)[1].lower()
    if file_ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail=f"File type not supported. Allowed: {', '.join(ALLOWED_EXTENSIONS)}"
        )

    # Read file content
    content = await file.read()

    # Validate file size
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(
            status_code=400,
            detail=f"File too large. Maximum size: {MAX_FILE_SIZE // 1024 // 1024}MB"
        )

    # Generate unique file ID
    file_id = str(uuid.uuid4())
    tenant_id = str(current_user.tenant_id)

    # Create upload directory if needed
    upload_path = os.path.join(UPLOAD_DIR, tenant_id)
    os.makedirs(upload_path, exist_ok=True)

    # Save file
    file_path = os.path.join(upload_path, f"{file_id}{file_ext}")
    with open(file_path, "wb") as f:
        f.write(content)

    logger.info(f"upload: UPLOAD_DIR={UPLOAD_DIR}")
    logger.info(f"upload: saved file to {file_path}")
    logger.info(f"upload: file_id={file_id}, tenant_id={tenant_id}")
    logger.info(f"upload: file exists after save={os.path.exists(file_path)}")

    return UploadResponse(
        file_id=file_id,
        file_name=file.filename,
        file_size=len(content),
        mime_type=file.content_type or "application/octet-stream",
        uploaded_at=datetime.utcnow(),
    )


@router.post("/parse/{file_id}", response_model=ParsedDocumentResponse)
async def parse_uploaded_document(
    file_id: str,
    document_type: Optional[DocumentTypeEnum] = None,
    current_user: User = Depends(get_current_user),
):
    """
    Parse an uploaded document and extract structured data.

    Args:
        file_id: The ID from the upload response
        document_type: Optional hint for document type

    Returns:
        Extracted data from the document
    """
    tenant_id = str(current_user.tenant_id)
    upload_path = os.path.join(UPLOAD_DIR, tenant_id)

    # Find the file
    file_path = None
    for ext in ALLOWED_EXTENSIONS:
        candidate = os.path.join(upload_path, f"{file_id}{ext}")
        if os.path.exists(candidate):
            file_path = candidate
            break

    if not file_path:
        raise HTTPException(status_code=404, detail="Uploaded file not found")

    file_ext = os.path.splitext(file_path)[1].lower()

    # Parse based on file type
    parser = DocumentParser()

    if file_ext == ".pdf":
        # Parse PDF
        text = await _extract_pdf_text(file_path)
        parsed = parser.parse_pdf_text(text)

    elif file_ext in [".xlsx", ".xls"]:
        # Parse Excel
        data = await _read_excel_file(file_path)
        parsed = parser.parse_excel_data(data)

    elif file_ext == ".csv":
        # Parse CSV
        data = await _read_csv_file(file_path)
        parsed = parser.parse_excel_data(data)

    else:
        raise HTTPException(status_code=400, detail="Unsupported file format")

    # Override document type if specified
    if document_type:
        parsed.document_type = document_type.value

    return ParsedDocumentResponse(
        document_type=parsed.document_type,
        confidence=parsed.confidence,
        invoice_no=parsed.invoice_no,
        invoice_date=parsed.invoice_date,
        bl_no=parsed.bl_no,
        bl_date=parsed.bl_date,
        seller_name=parsed.seller_name,
        seller_address=parsed.seller_address,
        consignee_name=parsed.consignee_name,
        consignee_address=parsed.consignee_address,
        vessel_name=parsed.vessel_name,
        voyage_no=parsed.voyage_no,
        flight_no=parsed.flight_no,
        loading_port=parsed.loading_port,
        discharge_port=parsed.discharge_port,
        eta=parsed.eta,
        total_packages=parsed.total_packages,
        gross_weight=parsed.gross_weight,
        net_weight=parsed.net_weight,
        volume=parsed.volume,
        container_numbers=parsed.container_numbers,
        currency=parsed.currency,
        total_value=parsed.total_value,
        incoterms=parsed.incoterms,
        items=[
            ParsedItemResponse(
                item_no=item.item_no,
                product_code=item.product_code,
                product_name=item.product_name,
                hs_code=item.hs_code,
                quantity=item.quantity,
                unit=item.unit,
                unit_price=item.unit_price,
                total_value=item.total_value,
                gross_weight=item.gross_weight,
                net_weight=item.net_weight,
                country_of_origin=item.country_of_origin,
            )
            for item in parsed.items
        ],
        warnings=parsed.warnings,
        extracted_fields=parsed.extracted_fields,
    )


@router.post("/parse-direct", response_model=ParsedDocumentResponse)
async def parse_document_direct(
    file: UploadFile = File(...),
    document_type: Optional[DocumentTypeEnum] = Form(None),
    current_user: User = Depends(get_current_user),
):
    """
    Upload and parse a document in one step.
    File is not saved - only parsed and returned.
    """
    # Validate file extension
    file_ext = os.path.splitext(file.filename)[1].lower()
    if file_ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail=f"File type not supported. Allowed: {', '.join(ALLOWED_EXTENSIONS)}"
        )

    # Read file content
    content = await file.read()

    # Validate file size
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(
            status_code=400,
            detail=f"File too large. Maximum size: {MAX_FILE_SIZE // 1024 // 1024}MB"
        )

    # Parse based on file type
    parser = DocumentParser()

    if file_ext == ".pdf":
        text = await _extract_pdf_text_from_bytes(content)
        parsed = parser.parse_pdf_text(text)

    elif file_ext in [".xlsx", ".xls"]:
        data = await _read_excel_from_bytes(content)
        parsed = parser.parse_excel_data(data)

    elif file_ext == ".csv":
        data = await _read_csv_from_bytes(content)
        parsed = parser.parse_excel_data(data)

    else:
        raise HTTPException(status_code=400, detail="Unsupported file format")

    # Override document type if specified
    if document_type:
        parsed.document_type = document_type.value

    return ParsedDocumentResponse(
        document_type=parsed.document_type,
        confidence=parsed.confidence,
        invoice_no=parsed.invoice_no,
        invoice_date=parsed.invoice_date,
        bl_no=parsed.bl_no,
        bl_date=parsed.bl_date,
        seller_name=parsed.seller_name,
        seller_address=parsed.seller_address,
        consignee_name=parsed.consignee_name,
        consignee_address=parsed.consignee_address,
        vessel_name=parsed.vessel_name,
        voyage_no=parsed.voyage_no,
        flight_no=parsed.flight_no,
        loading_port=parsed.loading_port,
        discharge_port=parsed.discharge_port,
        eta=parsed.eta,
        total_packages=parsed.total_packages,
        gross_weight=parsed.gross_weight,
        net_weight=parsed.net_weight,
        volume=parsed.volume,
        container_numbers=parsed.container_numbers,
        currency=parsed.currency,
        total_value=parsed.total_value,
        incoterms=parsed.incoterms,
        items=[
            ParsedItemResponse(
                item_no=item.item_no,
                product_code=item.product_code,
                product_name=item.product_name,
                hs_code=item.hs_code,
                quantity=item.quantity,
                unit=item.unit,
                unit_price=item.unit_price,
                total_value=item.total_value,
                gross_weight=item.gross_weight,
                net_weight=item.net_weight,
                country_of_origin=item.country_of_origin,
            )
            for item in parsed.items
        ],
        warnings=parsed.warnings,
        extracted_fields=parsed.extracted_fields,
    )


@router.post("/merge", response_model=CustomsDataPreview)
async def merge_parsed_documents(
    file_ids: List[str] = Form(...),
    current_user: User = Depends(get_current_user),
):
    """
    Merge multiple parsed documents into a single customs declaration preview.

    This combines data from Invoice, Packing List, B/L, Arrival Notice
    into a unified customs declaration structure.
    """
    tenant_id = str(current_user.tenant_id)
    upload_path = os.path.join(UPLOAD_DIR, tenant_id)

    parser = DocumentParser()
    parsed_docs: List[ParsedDocument] = []
    source_names: List[str] = []

    for file_id in file_ids:
        # Find and parse each file
        file_path = None
        for ext in ALLOWED_EXTENSIONS:
            candidate = os.path.join(upload_path, f"{file_id}{ext}")
            if os.path.exists(candidate):
                file_path = candidate
                break

        if not file_path:
            continue

        file_ext = os.path.splitext(file_path)[1].lower()

        if file_ext == ".pdf":
            text = await _extract_pdf_text(file_path)
            parsed = parser.parse_pdf_text(text)
        elif file_ext in [".xlsx", ".xls"]:
            data = await _read_excel_file(file_path)
            parsed = parser.parse_excel_data(data)
        elif file_ext == ".csv":
            data = await _read_csv_file(file_path)
            parsed = parser.parse_excel_data(data)
        else:
            continue

        parsed_docs.append(parsed)
        source_names.append(parsed.document_type or "UNKNOWN")

    if not parsed_docs:
        raise HTTPException(status_code=400, detail="No valid documents to merge")

    # Merge documents with priority rules
    merged = _merge_documents(parsed_docs)

    return CustomsDataPreview(
        declaration_type="IMPORT",
        foreign_partner_name=merged.seller_name,
        foreign_partner_address=merged.seller_address,
        trader_name=merged.consignee_name,
        trader_address=merged.consignee_address,
        invoice_no=merged.invoice_no,
        invoice_date=merged.invoice_date,
        bl_no=merged.bl_no,
        bl_date=merged.bl_date,
        vessel_name=merged.vessel_name,
        voyage_no=merged.voyage_no,
        loading_port=merged.loading_port,
        discharge_port=merged.discharge_port,
        total_packages=merged.total_packages,
        gross_weight=merged.gross_weight,
        net_weight=merged.net_weight,
        container_numbers=merged.container_numbers,
        currency_code=merged.currency,
        fob_value=merged.total_value,
        incoterms=merged.incoterms,
        items=[
            ParsedItemResponse(
                item_no=item.item_no,
                product_code=item.product_code,
                product_name=item.product_name,
                hs_code=item.hs_code,
                quantity=item.quantity,
                unit=item.unit,
                unit_price=item.unit_price,
                total_value=item.total_value,
                gross_weight=item.gross_weight,
                net_weight=item.net_weight,
                country_of_origin=item.country_of_origin,
            )
            for item in merged.items
        ],
        source_documents=source_names,
        data_confidence=merged.confidence,
    )


@router.delete("/upload/{file_id}")
async def delete_uploaded_file(
    file_id: str,
    current_user: User = Depends(get_current_user),
):
    """Delete an uploaded file."""
    tenant_id = str(current_user.tenant_id)
    upload_path = os.path.join(UPLOAD_DIR, tenant_id)

    # Find and delete the file
    deleted = False
    for ext in ALLOWED_EXTENSIONS:
        file_path = os.path.join(upload_path, f"{file_id}{ext}")
        if os.path.exists(file_path):
            os.remove(file_path)
            deleted = True
            break

    if not deleted:
        raise HTTPException(status_code=404, detail="File not found")

    return {"message": "File deleted"}


# Helper functions for file parsing

async def _extract_pdf_text(file_path: str) -> str:
    """Extract text from PDF file using pdfplumber."""
    try:
        import pdfplumber
        text_parts = []
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    text_parts.append(text)
        return "\n".join(text_parts)
    except ImportError:
        # Fallback to basic extraction if pdfplumber not available
        return await _extract_pdf_text_basic(file_path)


async def _extract_pdf_text_from_bytes(content: bytes) -> str:
    """Extract text from PDF bytes using pdfplumber."""
    try:
        import pdfplumber
        text_parts = []
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    text_parts.append(text)
        return "\n".join(text_parts)
    except ImportError:
        return await _extract_pdf_text_basic_bytes(content)


async def _extract_pdf_text_basic(file_path: str) -> str:
    """Basic PDF text extraction fallback."""
    try:
        import PyPDF2
        with open(file_path, "rb") as f:
            reader = PyPDF2.PdfReader(f)
            text_parts = []
            for page in reader.pages:
                text = page.extract_text()
                if text:
                    text_parts.append(text)
            return "\n".join(text_parts)
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to extract PDF text: {str(e)}. Please install pdfplumber or PyPDF2."
        )


async def _extract_pdf_text_basic_bytes(content: bytes) -> str:
    """Basic PDF text extraction from bytes."""
    try:
        import PyPDF2
        reader = PyPDF2.PdfReader(io.BytesIO(content))
        text_parts = []
        for page in reader.pages:
            text = page.extract_text()
            if text:
                text_parts.append(text)
        return "\n".join(text_parts)
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to extract PDF text: {str(e)}. Please install pdfplumber or PyPDF2."
        )


async def _read_excel_file(file_path: str) -> List[Dict[str, Any]]:
    """Read Excel file into list of dicts."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        # First row as headers
        headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows[0])]

        # Convert remaining rows to dicts
        data = []
        for row in rows[1:]:
            if any(cell is not None for cell in row):  # Skip empty rows
                row_dict = {headers[i]: cell for i, cell in enumerate(row) if i < len(headers)}
                data.append(row_dict)

        return data
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to read Excel: {str(e)}")


async def _read_excel_from_bytes(content: bytes) -> List[Dict[str, Any]]:
    """Read Excel from bytes into list of dicts."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows[0])]

        data = []
        for row in rows[1:]:
            if any(cell is not None for cell in row):
                row_dict = {headers[i]: cell for i, cell in enumerate(row) if i < len(headers)}
                data.append(row_dict)

        return data
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to read Excel: {str(e)}")


async def _read_csv_file(file_path: str) -> List[Dict[str, Any]]:
    """Read CSV file into list of dicts."""
    import csv
    with open(file_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        return list(reader)


async def _read_csv_from_bytes(content: bytes) -> List[Dict[str, Any]]:
    """Read CSV from bytes into list of dicts."""
    import csv
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return list(reader)


def _merge_documents(docs: List[ParsedDocument]) -> ParsedDocument:
    """
    Merge multiple parsed documents into one.

    Priority rules:
    - Invoice: Parties, values, items
    - B/L: Transport details, weights
    - Arrival Notice: ETA, warehouse
    - Packing List: Detailed items
    """
    result = ParsedDocument()
    result.confidence = 0.0

    # Track which document type provided each field
    for doc in docs:
        doc_type = doc.document_type

        # Invoice has priority for parties and values
        if doc_type == "INVOICE":
            result.seller_name = result.seller_name or doc.seller_name
            result.seller_address = result.seller_address or doc.seller_address
            result.consignee_name = result.consignee_name or doc.consignee_name
            result.consignee_address = result.consignee_address or doc.consignee_address
            result.invoice_no = result.invoice_no or doc.invoice_no
            result.invoice_date = result.invoice_date or doc.invoice_date
            result.currency = doc.currency if doc.currency != "USD" else result.currency
            result.total_value = doc.total_value if doc.total_value > 0 else result.total_value
            result.incoterms = result.incoterms or doc.incoterms

            # Use invoice items if available
            if doc.items and not result.items:
                result.items = doc.items

        # B/L has priority for transport
        elif doc_type == "BILL_OF_LADING":
            result.bl_no = result.bl_no or doc.bl_no
            result.bl_date = result.bl_date or doc.bl_date
            result.vessel_name = result.vessel_name or doc.vessel_name
            result.voyage_no = result.voyage_no or doc.voyage_no
            result.loading_port = result.loading_port or doc.loading_port
            result.discharge_port = result.discharge_port or doc.discharge_port
            result.gross_weight = doc.gross_weight if doc.gross_weight > 0 else result.gross_weight
            result.container_numbers = result.container_numbers or doc.container_numbers
            result.total_packages = doc.total_packages if doc.total_packages > 0 else result.total_packages

        # Airway Bill similar to B/L
        elif doc_type == "AIRWAY_BILL":
            result.bl_no = result.bl_no or doc.bl_no
            result.bl_date = result.bl_date or doc.bl_date
            result.flight_no = result.flight_no or doc.flight_no
            result.loading_port = result.loading_port or doc.loading_port
            result.discharge_port = result.discharge_port or doc.discharge_port
            result.gross_weight = doc.gross_weight if doc.gross_weight > 0 else result.gross_weight
            result.total_packages = doc.total_packages if doc.total_packages > 0 else result.total_packages

        # Arrival Notice supplements
        elif doc_type == "ARRIVAL_NOTICE":
            result.eta = result.eta or doc.eta
            result.vessel_name = result.vessel_name or doc.vessel_name
            result.loading_port = result.loading_port or doc.loading_port
            result.discharge_port = result.discharge_port or doc.discharge_port
            result.gross_weight = doc.gross_weight if doc.gross_weight > 0 and result.gross_weight == 0 else result.gross_weight
            result.volume = doc.volume if doc.volume > 0 else result.volume

        # Packing List for detailed items
        elif doc_type == "PACKING_LIST":
            # Merge items with HS code matching
            if doc.items:
                if not result.items:
                    result.items = doc.items
                else:
                    # Merge by matching product codes
                    _merge_items(result.items, doc.items)

            result.gross_weight = doc.gross_weight if doc.gross_weight > 0 else result.gross_weight
            result.net_weight = doc.net_weight if doc.net_weight > 0 else result.net_weight
            result.total_packages = doc.total_packages if doc.total_packages > 0 else result.total_packages

        # Fallback for any field
        else:
            for field in ["seller_name", "consignee_name", "bl_no", "vessel_name",
                          "loading_port", "discharge_port", "invoice_no"]:
                if not getattr(result, field) and getattr(doc, field):
                    setattr(result, field, getattr(doc, field))

        result.confidence = max(result.confidence, doc.confidence)

    return result


def _merge_items(existing: List[ParsedItem], new_items: List[ParsedItem]):
    """Merge item lists by matching product codes."""
    existing_codes = {item.product_code: item for item in existing if item.product_code}

    for new_item in new_items:
        if new_item.product_code and new_item.product_code in existing_codes:
            # Update existing item with new data
            existing_item = existing_codes[new_item.product_code]
            existing_item.gross_weight = new_item.gross_weight or existing_item.gross_weight
            existing_item.net_weight = new_item.net_weight or existing_item.net_weight
            existing_item.hs_code = new_item.hs_code or existing_item.hs_code
        else:
            # Add new item
            new_item.item_no = len(existing) + 1
            existing.append(new_item)


# ============================================================
# AI-POWERED PARSING ENDPOINTS
# ============================================================

class AIParseResponse(BaseModel):
    """Response schema for AI-parsed document"""
    success: bool
    error: Optional[str] = None
    confidence: float = 0.0
    provider_used: Optional[str] = None  # Which AI provider was used (gemini, claude, openai)

    # Document type
    document_type: Optional[str] = None

    # Header info
    invoice_no: Optional[str] = None
    invoice_date: Optional[str] = None
    bl_no: Optional[str] = None
    bl_date: Optional[str] = None

    # Parties - Exporter (foreign partner)
    exporter_name: Optional[str] = None
    exporter_address: Optional[str] = None
    exporter_country: Optional[str] = None

    # Parties - Importer
    importer_name: Optional[str] = None
    importer_address: Optional[str] = None
    importer_tax_code: Optional[str] = None

    # Transport
    vessel_name: Optional[str] = None
    voyage_no: Optional[str] = None
    flight_no: Optional[str] = None
    loading_port: Optional[str] = None
    loading_port_name: Optional[str] = None
    discharge_port: Optional[str] = None
    discharge_port_name: Optional[str] = None
    eta: Optional[str] = None

    # Cargo
    total_packages: int = 0
    package_unit: Optional[str] = None
    gross_weight: float = 0.0
    net_weight: float = 0.0
    volume_cbm: float = 0.0
    container_numbers: Optional[str] = None
    container_count: int = 0

    # Values
    currency: str = "USD"
    total_value: float = 0.0
    incoterms: Optional[str] = None

    # Items
    items: List[ParsedItemResponse] = []


@router.post("/parse-ai", response_model=AIParseResponse)
async def parse_document_with_ai_endpoint(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_session),
):
    """
    Parse a document using AI (Gemini/Claude/OpenAI).

    Sử dụng LLM để trích xuất dữ liệu từ PDF một cách thông minh,
    không phụ thuộc vào format cố định.

    Priority order được cấu hình trong AI Settings (Super Admin).
    Mặc định: Gemini Flash 1.5 (rẻ nhất), Claude (tốt nhất), OpenAI (backup).

    Hỗ trợ: Invoice, Packing List, Bill of Lading, Arrival Notice
    """
    # Validate file extension
    file_ext = os.path.splitext(file.filename)[1].lower()
    if file_ext != ".pdf":
        raise HTTPException(
            status_code=400,
            detail="AI parsing currently only supports PDF files"
        )

    # Read file content
    content = await file.read()

    # Validate file size
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(
            status_code=400,
            detail=f"File too large. Maximum size: {MAX_FILE_SIZE // 1024 // 1024}MB"
        )

    # Parse with AI - pass db session for config lookup
    result = await parse_document_with_ai(
        content,
        file.filename,
        db_session=db,
        tenant_id=getattr(current_user, 'tenant_id', None),
        user_id=str(current_user.id) if current_user else None,
    )

    # Convert to response
    return AIParseResponse(
        success=result.success,
        error=result.error,
        confidence=result.confidence,
        provider_used=result.provider_used,
        document_type=result.document_type,
        invoice_no=result.invoice_no,
        invoice_date=result.invoice_date,
        bl_no=result.bl_no,
        bl_date=result.bl_date,
        exporter_name=result.exporter_name,
        exporter_address=result.exporter_address,
        exporter_country=result.exporter_country,
        importer_name=result.importer_name,
        importer_address=result.importer_address,
        importer_tax_code=result.importer_tax_code,
        vessel_name=result.vessel_name,
        voyage_no=result.voyage_no,
        flight_no=result.flight_no,
        loading_port=result.loading_port,
        loading_port_name=result.loading_port_name,
        discharge_port=result.discharge_port,
        discharge_port_name=result.discharge_port_name,
        eta=result.eta,
        total_packages=result.total_packages,
        package_unit=result.package_unit,
        gross_weight=result.gross_weight,
        net_weight=result.net_weight,
        volume_cbm=result.volume_cbm,
        container_numbers=result.container_numbers,
        container_count=result.container_count,
        currency=result.currency,
        total_value=result.total_value,
        incoterms=result.incoterms,
        items=[
            ParsedItemResponse(
                item_no=item.get("item_no", idx + 1),
                product_code=item.get("product_code"),
                supplier_code=item.get("supplier_code"),
                product_name=item.get("product_name"),
                hs_code=item.get("hs_code"),
                quantity=item.get("quantity", 0),
                unit=item.get("unit"),
                unit_price=item.get("unit_price", 0),
                total_value=item.get("total_value", 0),
                gross_weight=item.get("gross_weight", 0),
                net_weight=item.get("net_weight", 0),
                country_of_origin=item.get("country_of_origin"),
            )
            for idx, item in enumerate(result.items)
        ],
    )


class AIBatchParseRequest(BaseModel):
    """Request to parse multiple files with AI"""
    file_ids: List[str]


class AIBatchParseResponse(BaseModel):
    """Response for batch AI parsing and merging"""
    success: bool
    error: Optional[str] = None
    confidence: float = 0.0

    # Merged declaration data
    declaration_type: str = "IMPORT"

    # Document numbers
    invoice_no: Optional[str] = None
    invoice_date: Optional[str] = None
    bl_no: Optional[str] = None
    bl_date: Optional[str] = None

    # Exporter (foreign partner)
    exporter_name: Optional[str] = None
    exporter_address: Optional[str] = None
    exporter_country: Optional[str] = None

    # Importer
    importer_name: Optional[str] = None
    importer_address: Optional[str] = None
    importer_tax_code: Optional[str] = None

    # Transport
    vessel_name: Optional[str] = None
    voyage_no: Optional[str] = None
    flight_no: Optional[str] = None
    loading_port: Optional[str] = None
    loading_port_name: Optional[str] = None
    discharge_port: Optional[str] = None
    discharge_port_name: Optional[str] = None
    eta: Optional[str] = None

    # Cargo
    total_packages: int = 0
    package_unit: Optional[str] = None
    gross_weight: float = 0.0
    net_weight: float = 0.0
    volume_cbm: float = 0.0
    container_numbers: Optional[str] = None
    container_count: int = 0

    # Values
    currency: str = "USD"
    total_value: float = 0.0
    incoterms: Optional[str] = None
    exchange_rate: float = 0.0

    # Items
    items: List[ParsedItemResponse] = []

    # Source info
    source_documents: List[str] = []

    # Field sources with original content (for tooltip display)
    # Key: field name, Value: source info including original content
    field_sources: Optional[dict] = None


@router.get("/ai-status")
async def check_ai_status(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_session),
):
    """
    Check AI API configuration status.
    Returns which AI providers are configured (from database).
    """
    from sqlalchemy import text

    # Query configured providers from database
    result = db.execute(text("""
        SELECT provider_code, is_configured, is_enabled
        FROM ai_providers
        WHERE is_configured = true AND api_key IS NOT NULL AND api_key != ''
    """))

    configured_providers = {}
    any_configured = False
    for row in result:
        provider_code = row[0]
        is_configured = row[1]
        is_enabled = row[2]
        configured_providers[provider_code] = {
            "configured": is_configured,
            "enabled": is_enabled
        }
        if is_configured:
            any_configured = True

    return {
        "gemini_configured": configured_providers.get("gemini", {}).get("configured", False),
        "claude_configured": configured_providers.get("claude", {}).get("configured", False),
        "openai_configured": configured_providers.get("openai", {}).get("configured", False),
        "any_configured": any_configured,
        "configured_providers": list(configured_providers.keys()),
        "upload_dir": UPLOAD_DIR,
        "upload_dir_exists": os.path.exists(UPLOAD_DIR),
    }


@router.post("/parse-ai-batch", response_model=AIBatchParseResponse)
async def parse_documents_batch_with_ai(
    file_ids: List[str] = Form(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_session),
):
    """
    Parse multiple uploaded PDF files with AI and merge results.

    Trích xuất dữ liệu từ nhiều file (Invoice, B/L, PKL, Arrival Notice)
    và gộp thành một tờ khai duy nhất.

    Priority order được cấu hình trong AI Settings (Super Admin).
    """
    import logging
    logger = logging.getLogger(__name__)

    try:
        tenant_id = str(current_user.tenant_id)
        upload_path = os.path.join(UPLOAD_DIR, tenant_id)

        logger.info(f"parse-ai-batch: UPLOAD_DIR={UPLOAD_DIR}")
        logger.info(f"parse-ai-batch: upload_path={upload_path}")
        logger.info(f"parse-ai-batch: file_ids={file_ids}")
        logger.info(f"parse-ai-batch: upload_path exists={os.path.exists(upload_path)}")

        # List files in directory for debugging
        if os.path.exists(upload_path):
            files_in_dir = os.listdir(upload_path)
            logger.info(f"parse-ai-batch: files in directory={files_in_dir}")
        else:
            logger.warning(f"parse-ai-batch: upload directory does not exist: {upload_path}")

        results: List[AIParseResult] = []
        source_docs: List[str] = []
        errors: List[str] = []

        for file_id in file_ids:
            # Find the file
            file_path = None
            for ext in ALLOWED_EXTENSIONS:
                candidate = os.path.join(upload_path, f"{file_id}{ext}")
                logger.info(f"parse-ai-batch: checking {candidate}, exists={os.path.exists(candidate)}")
                if os.path.exists(candidate):
                    file_path = candidate
                    break

            if not file_path:
                errors.append(f"File not found: {file_id}")
                logger.warning(f"parse-ai-batch: file not found for id={file_id}")
                continue

            if not file_path.endswith(".pdf"):
                errors.append(f"Not a PDF: {file_id}")
                logger.warning(f"parse-ai-batch: file is not PDF: {file_path}")
                continue

            # Read and parse
            logger.info(f"parse-ai-batch: parsing file {file_path}")
            with open(file_path, "rb") as f:
                content = f.read()

            result = await parse_document_with_ai(
                content,
                os.path.basename(file_path),
                db_session=db,
                tenant_id=tenant_id,
                user_id=str(current_user.id) if current_user else None,
            )
            if result.success:
                results.append(result)
                source_docs.append(result.document_type or "UNKNOWN")
                logger.info(f"parse-ai-batch: successfully parsed {file_id}, type={result.document_type}")
            else:
                errors.append(f"Parse failed: {file_id} - {result.error}")
                logger.warning(f"parse-ai-batch: AI parse failed for {file_id}: {result.error}")

        if not results:
            error_msg = "No valid PDF files to parse"
            if errors:
                error_msg = f"{error_msg}. Details: {'; '.join(errors)}"
            return AIBatchParseResponse(
                success=False,
                error=error_msg,
                source_documents=[]
            )

        # Merge results
        logger.info(f"parse-ai-batch: merging {len(results)} results")
        merged = _merge_ai_results(results)
        logger.info(f"parse-ai-batch: merged result - invoice={merged.invoice_no}, bl={merged.bl_no}, items={len(merged.items) if merged.items else 0}")

        response = AIBatchParseResponse(
            success=True,
            confidence=merged.confidence,
            declaration_type="IMPORT",
            invoice_no=merged.invoice_no,
            invoice_date=merged.invoice_date,
            bl_no=merged.bl_no,
            bl_date=merged.bl_date,
            exporter_name=merged.exporter_name,
            exporter_address=merged.exporter_address,
            exporter_country=merged.exporter_country,
            importer_name=merged.importer_name,
            importer_address=merged.importer_address,
            importer_tax_code=merged.importer_tax_code,
            vessel_name=merged.vessel_name,
            voyage_no=merged.voyage_no,
            flight_no=merged.flight_no,
            loading_port=merged.loading_port,
            loading_port_name=merged.loading_port_name,
            discharge_port=merged.discharge_port,
            discharge_port_name=merged.discharge_port_name,
            eta=merged.eta,
            total_packages=merged.total_packages,
            package_unit=merged.package_unit,
            gross_weight=merged.gross_weight,
            net_weight=merged.net_weight,
            volume_cbm=merged.volume_cbm,
            container_numbers=merged.container_numbers,
            container_count=merged.container_count,
            currency=merged.currency,
            total_value=merged.total_value,
            incoterms=merged.incoterms,
            exchange_rate=merged.exchange_rate,
            items=[
                ParsedItemResponse(
                    item_no=int(item.get("item_no", idx + 1) or idx + 1),
                    product_code=str(item.get("product_code") or "") or None,
                    supplier_code=str(item.get("supplier_code") or "") or None,
                    product_name=str(item.get("product_name") or "") or None,
                    hs_code=str(item.get("hs_code") or "") or None,
                    quantity=float(item.get("quantity") or 0),
                    unit=str(item.get("unit") or "") or None,
                    unit_price=float(item.get("unit_price") or 0),
                    total_value=float(item.get("total_value") or 0),
                    gross_weight=float(item.get("gross_weight") or 0),
                    net_weight=float(item.get("net_weight") or 0),
                    country_of_origin=str(item.get("country_of_origin") or "") or None,
                )
                for idx, item in enumerate(merged.items or [])
            ],
            source_documents=source_docs,
        )
        logger.info(f"parse-ai-batch: returning response with success={response.success}")
        return response
    except Exception as e:
        logger.exception(f"parse-ai-batch: Unexpected error: {e}")
        return AIBatchParseResponse(
            success=False,
            error=f"Server error: {str(e)}",
            source_documents=[]
        )


def _merge_ai_results(results: List[AIParseResult]) -> AIParseResult:
    """
    Merge multiple AI parse results into one.

    Priority rules:
    - Invoice: parties, values, items, incoterms
    - B/L: transport, weights, containers
    - Arrival Notice: ETA, ports
    - Packing List: detailed items, weights
    """
    merged = AIParseResult(success=True)

    for result in results:
        doc_type = result.document_type

        # Invoice has priority for parties and values
        if doc_type == "INVOICE":
            merged.invoice_no = merged.invoice_no or result.invoice_no
            merged.invoice_date = merged.invoice_date or result.invoice_date
            merged.exporter_name = merged.exporter_name or result.exporter_name
            merged.exporter_address = merged.exporter_address or result.exporter_address
            merged.exporter_country = merged.exporter_country or result.exporter_country
            merged.importer_name = merged.importer_name or result.importer_name
            merged.importer_address = merged.importer_address or result.importer_address
            merged.importer_tax_code = merged.importer_tax_code or result.importer_tax_code
            merged.currency = result.currency if result.currency != "USD" else merged.currency
            merged.total_value = result.total_value if result.total_value > 0 else merged.total_value
            merged.incoterms = merged.incoterms or result.incoterms
            if result.items and not merged.items:
                merged.items = result.items

        # B/L has priority for transport
        elif doc_type == "BILL_OF_LADING":
            merged.bl_no = merged.bl_no or result.bl_no
            merged.bl_date = merged.bl_date or result.bl_date
            merged.vessel_name = merged.vessel_name or result.vessel_name
            merged.voyage_no = merged.voyage_no or result.voyage_no
            merged.loading_port = merged.loading_port or result.loading_port
            merged.loading_port_name = merged.loading_port_name or result.loading_port_name
            merged.discharge_port = merged.discharge_port or result.discharge_port
            merged.discharge_port_name = merged.discharge_port_name or result.discharge_port_name
            merged.gross_weight = result.gross_weight if result.gross_weight > 0 else merged.gross_weight
            merged.container_numbers = merged.container_numbers or result.container_numbers
            merged.container_count = result.container_count if result.container_count > 0 else merged.container_count
            merged.total_packages = result.total_packages if result.total_packages > 0 else merged.total_packages
            merged.package_unit = merged.package_unit or result.package_unit

        # Arrival Notice supplements
        elif doc_type == "ARRIVAL_NOTICE":
            merged.eta = merged.eta or result.eta
            merged.vessel_name = merged.vessel_name or result.vessel_name
            merged.loading_port = merged.loading_port or result.loading_port
            merged.loading_port_name = merged.loading_port_name or result.loading_port_name
            merged.discharge_port = merged.discharge_port or result.discharge_port
            merged.discharge_port_name = merged.discharge_port_name or result.discharge_port_name
            merged.gross_weight = result.gross_weight if result.gross_weight > 0 and merged.gross_weight == 0 else merged.gross_weight
            merged.volume_cbm = result.volume_cbm if result.volume_cbm > 0 else merged.volume_cbm

        # Packing List for detailed items
        elif doc_type == "PACKING_LIST":
            merged.gross_weight = result.gross_weight if result.gross_weight > 0 else merged.gross_weight
            merged.net_weight = result.net_weight if result.net_weight > 0 else merged.net_weight
            merged.total_packages = result.total_packages if result.total_packages > 0 else merged.total_packages
            # Merge items
            if result.items:
                if not merged.items:
                    merged.items = result.items
                else:
                    _merge_ai_items(merged.items, result.items)

        # Fallback for any doc type
        else:
            for field in ["exporter_name", "importer_name", "bl_no", "vessel_name",
                          "loading_port", "discharge_port", "invoice_no"]:
                if not getattr(merged, field) and getattr(result, field):
                    setattr(merged, field, getattr(result, field))

        merged.confidence = max(merged.confidence, result.confidence)

    return merged


def _merge_ai_items(existing: List[Dict], new_items: List[Dict]):
    """Merge AI item lists by matching product codes."""
    existing_codes = {item.get("product_code"): item for item in existing if item.get("product_code")}

    for new_item in new_items:
        code = new_item.get("product_code")
        if code and code in existing_codes:
            # Update existing item with new data
            existing_item = existing_codes[code]
            for key in ["gross_weight", "net_weight", "hs_code", "supplier_code"]:
                if new_item.get(key) and not existing_item.get(key):
                    existing_item[key] = new_item[key]
        else:
            # Add new item
            new_item["item_no"] = len(existing) + 1
            existing.append(new_item)
