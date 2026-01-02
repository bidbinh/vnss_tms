"""
Create Excel template for fuel logs import
Run: python -m scripts.create_excel_template
"""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

def create_template():
    """Create Excel template with sample data"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fuel Logs"

    # Define headers
    headers = [
        "Ngày",
        "Số xe",
        "Tài xế",
        "Chỉ số đồng hồ Km xe",
        "Đổ thực tế",
        "Đơn giá",
        "Tổng tiền",
        "Ghi chú",
        "Trạng thái thanh toán"
    ]

    # Style for header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Write headers
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Add sample data
    sample_data = [
        ["2024-01-07", "50E-482.52", "Nguyễn Văn Tuyến", 129470, 250.67, 18750, 4700006, "Xe đổ dầu ngoài", "PAID"],
        ["2024-01-22", "50E-482.52", "Nguyễn Văn Tuyến", 131959, 252.82, 19780, 5000780, "Xe đổ dầu ngoài", "PAID"],
        ["2024-01-26", "50E-482.52", "Nguyễn Văn Tuyến", 132815, 235.38, 19070, 4500043, "Xe đổ dầu ngoài", "PAID"],
    ]

    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Adjust column widths
    column_widths = [12, 12, 18, 22, 12, 12, 12, 20, 20]
    for col, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # Save template
    output_path = Path(__file__).parent.parent / "fuel_logs_template.xlsx"
    wb.save(output_path)
    print(f"[OK] Template created: {output_path}")
    print("\nCột Excel:")
    print("1. Ngày - Format: YYYY-MM-DD (ví dụ: 2024-01-07)")
    print("2. Số xe - Biển số xe (ví dụ: 50E-482.52)")
    print("3. Tài xế - Tên tài xế (ví dụ: Nguyễn Văn Tuyến)")
    print("4. Chỉ số đồng hồ Km xe - Số km (ví dụ: 129470)")
    print("5. Đổ thực tế - Số lít (ví dụ: 250.67)")
    print("6. Đơn giá - Giá/lít VND (ví dụ: 18750)")
    print("7. Tổng tiền - Tổng số tiền VND (ví dụ: 4700006)")
    print("8. Ghi chú - Tùy chọn (ví dụ: Xe đổ dầu ngoài)")
    print("9. Trạng thái thanh toán - PAID hoặc UNPAID (mặc định: UNPAID)")

if __name__ == "__main__":
    create_template()
