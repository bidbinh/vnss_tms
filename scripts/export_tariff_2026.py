"""
Export tariff data from Excel BT2026 sheet to JSON
Usage: python export_tariff_2026.py
"""
import json
import openpyxl
import sys
import os

# Column mapping based on Excel structure (1-indexed)
COLUMNS = {
    'level': 5,      # V - level/indent
    'hs': 6,         # Ma hang
    'vi': 7,         # Mo ta tieng Viet
    'en': 8,         # Mo ta tieng Anh
    'unit_vi': 9,    # Don vi tinh
    'unit_en': 10,   # Unit of quantity

    # Import tax rates
    'nk_tt': 11,     # NK TT (Import - Normal)
    'nk_ud': 14,     # NK uu dai (Import - Preferential)
    'vat': 17,       # VAT

    # FTA rates
    'acfta': 20,     # ASEAN-China FTA
    'atiga': 23,     # ASEAN Trade in Goods Agreement
    'ajcep': 26,     # ASEAN-Japan CEP
    'vjepa': 29,     # Vietnam-Japan EPA
    'akfta': 32,     # ASEAN-Korea FTA
    'aanzfta': 35,   # ASEAN-Australia-New Zealand FTA
    'aifta': 38,     # ASEAN-India FTA
    'vkfta': 41,     # Vietnam-Korea FTA
    'vcfta': 44,     # Vietnam-Chile FTA
    'vneaeu': 47,    # Vietnam-Eurasian Economic Union
    'cptpp': 50,     # CPTPP
    'ahkfta': 53,    # ASEAN-Hong Kong FTA
    'vncu': 56,      # Vietnam-Cuba
    'evfta': 59,     # EU-Vietnam FTA
    'ukvfta': 62,    # UK-Vietnam FTA
    'vnlao': 65,     # Vietnam-Laos
    'vifta': 68,     # Vietnam-Israel FTA

    # RCEPT columns (71-78)
    'rcept_a': 71,
    'rcept_b': 72,
    'rcept_c': 73,
    'rcept_d': 74,
    'rcept_e': 75,
    'rcept_f': 76,

    # Other taxes
    'ttdb': 82,      # Thue tieu thu dac biet
    'xk': 85,        # Thue xuat khau
    'xk_cptpp': 88,  # XK CPTPP
    'xk_ev': 91,     # XK EVFTA
    'xk_ukv': 94,    # XK UKVFTA
    'bvmt': 97,      # Thue bao ve moi truong

    # Policy
    'policy': 100,   # Chinh sach mat hang theo ma HS
    'vat_reduce': 101,  # Giam VAT
}

def safe_str(val):
    """Convert value to string safely"""
    if val is None:
        return ""
    return str(val).strip()

def clean_rate(val):
    """Clean tax rate value - keep as string for special values like */5/8/10"""
    if val is None:
        return ""
    s = str(val).strip()
    if not s or s.lower() in ['none', 'nan', '-']:
        return ""
    return s

def export_tariff(excel_path, output_path, start_row=20):
    """Export tariff data from Excel to JSON"""
    print(f"Loading Excel file: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb['BT2026']

    # Find max row with data
    max_row = ws.max_row
    print(f"Total rows: {max_row}")

    data = []
    row_count = 0

    for row in range(start_row, max_row + 1):
        # Get basic info
        level = ws.cell(row=row, column=COLUMNS['level']).value
        hs = safe_str(ws.cell(row=row, column=COLUMNS['hs']).value)
        vi = safe_str(ws.cell(row=row, column=COLUMNS['vi']).value)
        en = safe_str(ws.cell(row=row, column=COLUMNS['en']).value)

        # Skip empty rows
        if not vi and not hs:
            continue

        # Skip section headers that don't have HS codes or tax info
        # (like "PHAN I", "Chu giai", etc.)
        if not hs and level is None and not vi.startswith('-'):
            continue

        item = {
            'l': int(level) if level is not None else 0,
            'hs': hs,
            'vi': vi,
            'en': en,
            'u': safe_str(ws.cell(row=row, column=COLUMNS['unit_vi']).value),
        }

        # Import tax rates
        item['nk_tt'] = clean_rate(ws.cell(row=row, column=COLUMNS['nk_tt']).value)
        item['nk_ud'] = clean_rate(ws.cell(row=row, column=COLUMNS['nk_ud']).value)
        item['vat'] = clean_rate(ws.cell(row=row, column=COLUMNS['vat']).value)

        # FTA rates
        item['acfta'] = clean_rate(ws.cell(row=row, column=COLUMNS['acfta']).value)
        item['atiga'] = clean_rate(ws.cell(row=row, column=COLUMNS['atiga']).value)
        item['ajcep'] = clean_rate(ws.cell(row=row, column=COLUMNS['ajcep']).value)
        item['vjepa'] = clean_rate(ws.cell(row=row, column=COLUMNS['vjepa']).value)
        item['akfta'] = clean_rate(ws.cell(row=row, column=COLUMNS['akfta']).value)
        item['aanzfta'] = clean_rate(ws.cell(row=row, column=COLUMNS['aanzfta']).value)
        item['aifta'] = clean_rate(ws.cell(row=row, column=COLUMNS['aifta']).value)
        item['vkfta'] = clean_rate(ws.cell(row=row, column=COLUMNS['vkfta']).value)
        item['vcfta'] = clean_rate(ws.cell(row=row, column=COLUMNS['vcfta']).value)
        item['vneaeu'] = clean_rate(ws.cell(row=row, column=COLUMNS['vneaeu']).value)
        item['cptpp'] = clean_rate(ws.cell(row=row, column=COLUMNS['cptpp']).value)
        item['ahkfta'] = clean_rate(ws.cell(row=row, column=COLUMNS['ahkfta']).value)
        item['vncu'] = clean_rate(ws.cell(row=row, column=COLUMNS['vncu']).value)
        item['evfta'] = clean_rate(ws.cell(row=row, column=COLUMNS['evfta']).value)
        item['ukvfta'] = clean_rate(ws.cell(row=row, column=COLUMNS['ukvfta']).value)
        item['vnlao'] = clean_rate(ws.cell(row=row, column=COLUMNS['vnlao']).value)
        item['vifta'] = clean_rate(ws.cell(row=row, column=COLUMNS['vifta']).value)

        # RCEPT
        item['rcept_a'] = clean_rate(ws.cell(row=row, column=COLUMNS['rcept_a']).value)
        item['rcept_b'] = clean_rate(ws.cell(row=row, column=COLUMNS['rcept_b']).value)
        item['rcept_c'] = clean_rate(ws.cell(row=row, column=COLUMNS['rcept_c']).value)
        item['rcept_d'] = clean_rate(ws.cell(row=row, column=COLUMNS['rcept_d']).value)
        item['rcept_e'] = clean_rate(ws.cell(row=row, column=COLUMNS['rcept_e']).value)
        item['rcept_f'] = clean_rate(ws.cell(row=row, column=COLUMNS['rcept_f']).value)

        # Other taxes
        item['ttdb'] = clean_rate(ws.cell(row=row, column=COLUMNS['ttdb']).value)
        item['xk'] = clean_rate(ws.cell(row=row, column=COLUMNS['xk']).value)
        item['xk_cptpp'] = clean_rate(ws.cell(row=row, column=COLUMNS['xk_cptpp']).value)
        item['xk_ev'] = clean_rate(ws.cell(row=row, column=COLUMNS['xk_ev']).value)
        item['xk_ukv'] = clean_rate(ws.cell(row=row, column=COLUMNS['xk_ukv']).value)
        item['bvmt'] = clean_rate(ws.cell(row=row, column=COLUMNS['bvmt']).value)

        # Policy
        item['policy'] = safe_str(ws.cell(row=row, column=COLUMNS['policy']).value)
        item['vat_reduce'] = safe_str(ws.cell(row=row, column=COLUMNS['vat_reduce']).value)

        data.append(item)
        row_count += 1

        if row_count % 1000 == 0:
            print(f"Processed {row_count} rows...")

    print(f"Total items exported: {len(data)}")

    # Write to JSON
    print(f"Writing to: {output_path}")
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, separators=(',', ':'))

    print("Done!")
    return data

if __name__ == '__main__':
    # Get script directory
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(script_dir)

    excel_path = os.path.join(project_root, 'BIEU THUE XNK 2026.01.xlsx')
    output_path = os.path.join(project_root, 'frontend', 'public', 'data', 'tariff-2026.json')

    if not os.path.exists(excel_path):
        print(f"Error: Excel file not found: {excel_path}")
        sys.exit(1)

    export_tariff(excel_path, output_path)
